from tkinter import *
import tkinter as tk
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

File_Path = "Backend_Data.xlsx"

root = Tk()
root.title("Data Entry")
root.geometry("750x450+300+200")
root.resizable(True,False)
root.configure(bg="#326273")

File = pathlib.Path(File_Path)
if File.exists():
    pass
else:
    File = Workbook()
    Sheet = File.active
    Sheet['A1'] = "Full Name"
    Sheet['B1'] = "Phone Number"
    Sheet['C1'] = "Age"
    Sheet['D1'] = "Gender"
    Sheet['E1'] = "Address"

    File.save(File_Path)


def clear():
    NameValue.set('')
    ContactValue.set('')
    AgeValue.set('')
    AddressEntry.delete(1.0,END)

def submit():
    Name = NameValue.get()
    Contact = ContactValue.get()
    Age = AgeValue.get()
    Gender = GenderCombobox.get()
    Address = AddressEntry.get(1.0,END)

    File = openpyxl.load_workbook(File_Path)
    Sheet = File.active
    Sheet.cell(column=1,row=Sheet.max_row+1,value=Name)
    Sheet.cell(column=2,row=Sheet.max_row,value=Contact)
    Sheet.cell(column=3,row=Sheet.max_row,value=Age)
    Sheet.cell(column=4,row=Sheet.max_row,value=Gender)
    Sheet.cell(column=5,row=Sheet.max_row,value=Address)
    
    File.save(r'Backend_Data.xlsx')

    messagebox.showinfo("info","Details Added Successfully")
    clear()

#icon
# Icon = PhotoImage(file="Logo.png")
# root.iconphoto(False,Icon)

#heading
Label(root,text="Please fill out this Entry from:",font="arial 13",bg="#326273",fg="#fff").place(x=20,y=20)

#label
Label(root,text="Name",font=23,bg="#326273",fg="#fff").place(x=50,y=100)
Label(root,text="Contancr no.",font=23,bg="#326273",fg="#fff").place(x=50,y=150)
Label(root,text="Age",font=23,bg="#326273",fg="#fff").place(x=50,y=200)
Label(root,text="Gender",font=23,bg="#326273",fg="#fff").place(x=385,y=200)
Label(root,text="Address",font=23,bg="#326273",fg="#fff").place(x=50,y=250)

#entry
NameValue = StringVar()
ContactValue = IntVar()
AgeValue = IntVar()

NameEntry = Entry(root,textvariable=NameValue,width=37,bd=2,font=20)
ContactEntry = Entry(root,textvariable=ContactValue,width=37,bd=2,font=20)
AgeEntry = Entry(root,textvariable=AgeValue,width=15,bd=2,font=20)

NameEntry.place(x=200,y=100)
ContactEntry.place(x=200,y=150)
AgeEntry.place(x=200,y=200)

#gender
GenderCombobox = Combobox(root,values=['Male','Female'],font="arial 14",state='r',width=12)
GenderCombobox.place(x=460,y=200)
GenderCombobox.set('Choose')

AddressEntry = Text(root,width=50,height=4,bd=4)
AddressEntry.place(x=200,y=250)

Button(root,text="Submit",bg="#326273",fg="white",width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Clear",bg="#326273",fg="white",width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit",bg="#326273",fg="white",width=15,height=2,command=lambda: root.destroy()).place(x=480,y=350)



root.mainloop()